
import requests
import re
from openpyxl import Workbook
import time

start_time=time.time()
hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'}, \
       {
           'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'}, \
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]
tags = ['小说', '生活']

def book_spider(tags):
    wb = Workbook()
    ws = []
    for i  in range(len(tags)):
        ws.append(wb.create_sheet(title=tags[i]))
        for page_num in range(1,51):
            url = 'http://www.douban.com/tag/' + str(tags[i]) + '/book?start=' + str(page_num * 20)+'0&type=T'
            try:
                res = requests.get(url, headers=hds[page_num % 3])
                if res.status_code == 200:
                    print('请求成功')

                    pattern = re.compile(
                        '<li.*?subject-item.*?img.*?src="(.*?)".*?title="(.*?)".*?</a>.*?pub.*?>(.*?)</div>' +
                        '.*?rating_nums.*?>(.*?)</span>.*?pl.*?>(.*?)</span>' +
                        '.*?<p>(.*?)</p>.*?</li>', re.S)
                    items = re.findall(pattern, res.text)
                    count = 1
                    for item in items:
                        item = re.sub('\\\\n', '', str(item))
                        item = re.sub("'", "", item)
                        item = re.sub(' ', '', item)
                        print(item)
                        ws[i].append([count, item])
                        count += 1
                else:print('请求失败')

            except:
                raise Exception
            time.sleep(1)
    wb.save(filename='E:/Desktop/12.26豆瓣.xlsx')
    print( 'total time=%.2f' %(time.time()-start_time))

if __name__=='__main__':
    book_spider(tags)




